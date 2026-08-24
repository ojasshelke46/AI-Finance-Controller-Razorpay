"""Parse bank CSV and ledger XLSX source files back into txns.

Handles the mess these files actually arrive in: junk rows above the
real header, two date formats, rupee strings with Indian comma grouping
and currency symbols, CR/DR markers, and references buried inside
free-text bank narration.

Amounts go through lib.money, so a rupee string becomes integer paise
without a float ever being constructed.
"""

import csv
import logging
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from lib import db
from lib.money import rupee_string_to_paise

logger = logging.getLogger("ingest.files")

# Banking boilerplate that shows up around the reference in narration.
_NARRATION_STOPWORDS = {
    "NEFT", "IMPS", "UPI", "RTGS", "TRF", "FROM", "TO", "BY", "TRANSFER",
    "RAZORPAY", "RZPY", "MPS", "SOFTWARE", "PVT", "LTD", "PRIVATE", "LIMITED",
    "SETTLEMENT", "PAYMENT", "COLLECTION", "REF", "REFNO", "DT", "DATE",
    "CR", "DR", "TXN", "NA",
}

# Narration splits on separators banks use; underscore is NOT a separator
# because references themselves contain it (pay_, setl_, rfnd_).
_NARRATION_SPLIT_RE = re.compile(r"[\s/\-*|:;,#()\[\]]+")
# Bare 1-2 digit fragments are date pieces ("DT 09/07"), not references.
_DATE_FRAGMENT_RE = re.compile(r"^\d{1,2}$")

_HEADER_TOKENS = {"date", "voucher no", "particulars", "amount"}

_INSERT_CHUNK_SIZE = 500


class FileParseError(RuntimeError):
    """Raised when a source file can't be parsed."""


def extract_reference(narration: str) -> str | None:
    """Pull the transaction reference out of free-text bank narration.

    'BY TRANSFER-NEFT*rfnd_QVmvb0*RAZORPAY SOFTWARE' -> 'rfnd_QVmvb0'
    'TRF FROM RAZORPAY REF 0lEDmC DT 09/07'          -> '0lEDmC'
    """
    if not narration:
        return None
    candidates = []
    for token in _NARRATION_SPLIT_RE.split(narration.strip()):
        if not token:
            continue
        if token.upper() in _NARRATION_STOPWORDS:
            continue
        if _DATE_FRAGMENT_RE.match(token):
            continue
        candidates.append(token)
    if not candidates:
        return None
    # If several survive, the reference is the longest — boilerplate that
    # escaped the stopword list is almost always shorter.
    return max(candidates, key=len)


def parse_date(text: str | None) -> str | None:
    """Accept DD/MM/YYYY, DD-MM-YYYY, and ISO. Returns ISO or None."""
    if text is None:
        return None
    if isinstance(text, (datetime, date)):
        d = text.date() if isinstance(text, datetime) else text
        return d.isoformat()
    s = str(text).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    raise FileParseError(f"unrecognised date format: {text!r}")


def parse_bank_csv(path: Path, batch_id: str, *, skipped: list[dict] | None = None) -> list[dict]:
    """Parse the bank statement. debit -> negative paise, credit ->
    positive. The running balance column is carried into raw but not
    used as an amount.

    A row this cannot parse is appended to `skipped` with its line
    number and the reason, and passed over, rather than aborting the
    whole file. One corrupt line in a bank export must not cost the
    other two thousand good rows — and a silently dropped row would be
    worse still, so every skip is returned to the caller and written to
    audit_log.
    """
    rows: list[dict] = []
    skipped = skipped if skipped is not None else []

    def _skip(lineno: int, reason: str, record: dict) -> None:
        logger.warning("line %d: %s, skipping", lineno, reason)
        skipped.append({
            "line": lineno,
            "reason": reason,
            "record": {k: v for k, v in record.items() if k},
        })

    with Path(path).open(newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for lineno, record in enumerate(reader, start=2):
            narration = (record.get("narration") or "").strip()
            debit = (record.get("debit") or "").strip()
            credit = (record.get("credit") or "").strip()

            if not debit and not credit:
                _skip(lineno, "no debit or credit value", record)
                continue
            if debit and credit:
                _skip(lineno, "both debit and credit populated", record)
                continue

            try:
                amount = -rupee_string_to_paise(debit) if debit else rupee_string_to_paise(credit)
            except (ValueError, TypeError) as exc:
                _skip(lineno, f"unparseable amount: {exc}", record)
                continue

            try:
                txn_date = parse_date(record.get("value_date"))
            except FileParseError as exc:
                _skip(lineno, f"unparseable date: {exc}", record)
                continue

            rows.append({
                "batch_id": batch_id,
                "source_kind": "bank",
                "external_ref": extract_reference(narration),
                "amount_paise": amount,
                "fee_paise": None,
                "tax_paise": None,
                "net_paise": None,
                "txn_date": txn_date,
                "value_date": txn_date,
                "description": narration,
                "counterparty": None,
                "raw": {"source_file": Path(path).name, "line": lineno, **record},
            })
    return rows


def _find_header_row(ws) -> int:
    """Locate the real header row beneath the junk block."""
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        values = {str(c).strip().lower() for c in row if c is not None}
        if len(_HEADER_TOKENS & values) >= 3:
            return idx
    raise FileParseError(f"no header row found in sheet {ws.title!r}")


def parse_ledger_xlsx(path: Path, batch_id: str, *, skipped: list[dict] | None = None) -> list[dict]:
    """Parse every sheet of the ledger workbook, skipping each sheet's
    junk header block. Amounts carry a trailing CR/DR marker.

    Unparseable rows are recorded in `skipped` and passed over, same
    contract as parse_bank_csv."""
    wb = load_workbook(Path(path), read_only=True, data_only=True)
    rows: list[dict] = []
    skipped = skipped if skipped is not None else []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = _find_header_row(ws)
        header = [str(c).strip().lower() if c is not None else ""
                  for c in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
        col = {name: i for i, name in enumerate(header)}

        for excel_row, values in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
        ):
            if values is None or all(v is None or str(v).strip() == "" for v in values):
                continue

            def cell(name: str):
                idx = col.get(name)
                return values[idx] if idx is not None and idx < len(values) else None

            amount_text = cell("amount")
            if amount_text is None or str(amount_text).strip() == "":
                logger.warning("%s row %d: no amount, skipping", sheet_name, excel_row)
                skipped.append({"sheet": sheet_name, "line": excel_row,
                                "reason": "no amount value", "record": {}})
                continue

            # rupee_string_to_paise reads the trailing CR/DR marker itself.
            try:
                amount = rupee_string_to_paise(str(amount_text))
            except (ValueError, TypeError) as exc:
                logger.warning("%s row %d: unparseable amount (%s), skipping",
                               sheet_name, excel_row, exc)
                skipped.append({"sheet": sheet_name, "line": excel_row,
                                "reason": f"unparseable amount: {exc}",
                                "record": {"amount": str(amount_text)}})
                continue

            rows.append({
                "batch_id": batch_id,
                "source_kind": "ledger",
                "external_ref": str(cell("voucher no")).strip() if cell("voucher no") else None,
                "amount_paise": amount,
                "fee_paise": None,
                "tax_paise": None,
                "net_paise": None,
                "txn_date": parse_date(cell("date")),
                "value_date": None,
                "description": str(cell("particulars") or "").strip(),
                "counterparty": None,
                "raw": {
                    "source_file": Path(path).name,
                    "sheet": sheet_name,
                    "row": excel_row,
                    "date": str(cell("date") or ""),
                    "voucher_no": str(cell("voucher no") or ""),
                    "particulars": str(cell("particulars") or ""),
                    "amount": str(amount_text),
                },
            })

    wb.close()
    return rows


def _insert(rows: list[dict]) -> None:
    for i in range(0, len(rows), _INSERT_CHUNK_SIZE):
        chunk = rows[i:i + _INSERT_CHUNK_SIZE]
        db.run_with_retry(lambda c=chunk: db.get_client().table("txns").insert(c).execute())


def _write_audit(batch_id: str, action: str, detail: dict) -> None:
    db.run_with_retry(
        lambda: db.get_client()
        .table("audit_log")
        .insert({
            "batch_id": batch_id,
            "actor": "scheduler",
            "step": "ingest_files",
            "action": action,
            "detail": detail,
        })
        .execute()
    )


def ingest_files(batch_id: str, bank_csv: Path, ledger_xlsx: Path) -> dict:
    """Parse both source files into txns for the given batch."""
    _write_audit(batch_id, "start", {"bank_csv": str(bank_csv), "ledger_xlsx": str(ledger_xlsx)})

    bank_skipped: list[dict] = []
    ledger_skipped: list[dict] = []
    bank_rows = parse_bank_csv(bank_csv, batch_id, skipped=bank_skipped)
    ledger_rows = parse_ledger_xlsx(ledger_xlsx, batch_id, skipped=ledger_skipped)

    _insert(bank_rows)
    _insert(ledger_rows)

    # Skipped rows get their own audit entry, not just a count in the
    # summary: "we dropped 3 rows" is not actionable, "we dropped line
    # 47 because both debit and credit were populated" is.
    if bank_skipped or ledger_skipped:
        _write_audit(batch_id, "records_skipped", {
            "bank_skipped_count": len(bank_skipped),
            "ledger_skipped_count": len(ledger_skipped),
            "bank_skipped": bank_skipped[:50],
            "ledger_skipped": ledger_skipped[:50],
            "truncated": len(bank_skipped) > 50 or len(ledger_skipped) > 50,
        })

    summary = {
        "bank_rows_parsed": len(bank_rows),
        "ledger_rows_parsed": len(ledger_rows),
        "bank_rows_skipped": len(bank_skipped),
        "ledger_rows_skipped": len(ledger_skipped),
        "bank_missing_ref": sum(1 for r in bank_rows if not r["external_ref"]),
        "ledger_missing_ref": sum(1 for r in ledger_rows if not r["external_ref"]),
        "total_rows": len(bank_rows) + len(ledger_rows),
        "skipped_lines": [s["line"] for s in bank_skipped + ledger_skipped][:50],
    }
    _write_audit(batch_id, "finish", summary)
    return summary
